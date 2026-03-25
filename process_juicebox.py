#!/usr/bin/env python3
"""
HeroScouter - Juicebox JSON to Excel Converter
Usage: python3 process_juicebox.py input.json [output.xlsx]
"""
import json, sys, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def extract(data):
    rows = []
    contacts = data if isinstance(data, list) else data.get('contacts', [data] if isinstance(data, dict) else [])
    
    for c in contacts:
        p = c.get('profileData', c)
        exp = p.get('experience', c.get('experience', []))
        edu = p.get('education', c.get('education', []))
        auto = p.get('autopilotData', c.get('autopilotData', {})) or {}
        avail = p.get('contact_info_availability', c.get('contact_info_availability', {})) or {}

        curr_title = exp[0].get('title', {}).get('name', '') if exp else ''
        curr_company = exp[0].get('company', {}).get('name', '') if exp else ''
        prev_companies = ' | '.join([e.get('company', {}).get('name', '') for e in exp[1:6] if e.get('company', {}).get('name')])
        prev_titles = ' | '.join([e.get('title', {}).get('name', '') for e in exp[1:6] if e.get('title', {}).get('name')])

        school = ''
        degree = ''
        if edu:
            school = edu[0].get('school', {}).get('name', '') or edu[0].get('name', '')
            degs = edu[0].get('degrees', [])
            degree = degs[0] if degs else ''

        crit = {}
        for r in auto.get('responses', []):
            t = r.get('criteriaText', '')
            v = r.get('result', '')
            if '1' in t and ('year' in t.lower() or 'yr' in t.lower()):
                crit['sdr'] = v
            elif 'SaaS' in t or 'FinTech' in t:
                crit['saas'] = v
            elif 'enterprise' in t.lower() or 'mid-market' in t.lower():
                crit['ent'] = v
            elif 'InsurTech' in t or 'AI' in t:
                crit['ai'] = v

        linkedin = p.get('linkedin_url', c.get('linkedin_url', ''))
        if linkedin and not linkedin.startswith('http'):
            linkedin = 'https://' + linkedin

        skills = p.get('skills', c.get('skills', []))
        summary = (p.get('summary', '') or '')[:250]

        rows.append({
            'Full Name': (p.get('full_name', c.get('full_name', '')) or '').title(),
            'First Name': (p.get('first_name', c.get('first_name', '')) or '').title(),
            'Last Name': (p.get('last_name', c.get('last_name', '')) or '').title(),
            'Current Title': curr_title,
            'Current Company': curr_company,
            'Location': p.get('location_name', c.get('location_name', '')),
            'LinkedIn URL': linkedin,
            'Match Rate %': auto.get('matchRateRounded', ''),
            'Work Email': 'Yes' if avail.get('work_email') else 'No',
            'Personal Email': 'Yes' if avail.get('personal_emails') else 'No',
            'Phone': 'Yes' if avail.get('phone_numbers') else 'No',
            'Prev Companies': prev_companies,
            'Prev Titles': prev_titles,
            'School': school,
            'Degree': degree,
            'Top Skills': ', '.join(skills[:12]),
            'SDR 1-3yr': crit.get('sdr', ''),
            'SaaS/FinTech': crit.get('saas', ''),
            'Enterprise Outbound': crit.get('ent', ''),
            'InsurTech/AI': crit.get('ai', ''),
            'Summary': summary,
            'Personalization Note': '',  # blank col for user to fill outreach notes
        })
    return rows

def make_xlsx(rows, outpath):
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidates"

    headers = list(rows[0].keys())
    
    H_FILL = PatternFill("solid", start_color="1F4E79")
    H_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    ALT_FILL = PatternFill("solid", start_color="EBF3FB")
    G = PatternFill("solid", start_color="C6EFCE")
    Y = PatternFill("solid", start_color="FFEB9C")
    R = PatternFill("solid", start_color="FFC7CE")
    DF = Font(name='Arial', size=9)

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(1, ci, h)
        cell.font = H_FONT; cell.fill = H_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 38

    mr_col = headers.index('Match Rate %') + 1
    crit_cols = {headers.index(h)+1 for h in ['SDR 1-3yr','SaaS/FinTech','Enterprise Outbound','InsurTech/AI']}

    for ri, row in enumerate(rows, 2):
        alt = (ri % 2 == 0)
        for ci, h in enumerate(headers, 1):
            v = row.get(h, '')
            cell = ws.cell(ri, ci, v)
            cell.font = DF
            if alt: cell.fill = ALT_FILL
            cell.alignment = Alignment(horizontal='left' if ci <= 7 else 'center', vertical='center')
            
            if ci == mr_col and v != '':
                try: cell.fill = G if float(v)>=90 else (Y if float(v)>=80 else R)
                except: pass
            if ci in crit_cols and v:
                cell.fill = G if v=='yes' else (Y if v=='maybe' else R if v=='no' else cell.fill)
        ws.row_dimensions[ri].height = 16

    widths = {'Full Name':22,'First Name':14,'Last Name':14,'Current Title':32,'Current Company':22,
              'Location':24,'LinkedIn URL':42,'Match Rate %':12,'Work Email':12,'Personal Email':13,
              'Phone':9,'Prev Companies':44,'Prev Titles':44,'School':28,'Degree':20,
              'Top Skills':45,'SDR 1-3yr':12,'SaaS/FinTech':13,'Enterprise Outbound':18,'InsurTech/AI':13,
              'Summary':50,'Personalization Note':35}
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(h, 15)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    wb.save(outpath)
    print(f"✓ Saved {len(rows)} candidates → {outpath}")

if __name__ == '__main__':
    inp = sys.argv[1] if len(sys.argv) > 1 else None
    out = sys.argv[2] if len(sys.argv) > 2 else 'candidates.xlsx'
    if inp:
        with open(inp) as f: data = json.load(f)
    else:
        data = json.load(sys.stdin)
    rows = extract(data)
    make_xlsx(rows, out)
